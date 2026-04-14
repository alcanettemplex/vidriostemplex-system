
import urllib.request, json
from openpyxl import load_workbook
from collections import defaultdict

SUPABASE_URL = 'https://mvcmofhbjvxjnxefylsa.supabase.co'
ANON_KEY = 'sb_publishable_4P_mqyw7E_p2yPHuo8hCRA_BgNR0Wsm'

# --- Leer Excel ---
wb = load_workbook(r'C:\Users\User\Desktop\vidrios-templex-system\documentation\INVENTARIO PERFILERIA ACTUAL 2026.xlsx', data_only=True)
ws = wb['INV']
rows = list(ws.iter_rows(min_row=2, values_only=True))
filled = [r for r in rows if r[1] is not None]

# Agrupar por codigo (multiples barras por codigo)
por_codigo = defaultdict(list)
for r in filled:
    cons, cod, desc, color, ref, mm, ubic = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
    por_codigo[cod].append({
        'cons': cons,
        'desc': str(desc) if desc else '',
        'color': str(color) if color else '',
        'ref': str(ref) if ref else '',
        'mm': mm or 0,
        'ubic': str(ubic) if ubic else ''
    })

# --- Consultar Supabase ---
url = f'{SUPABASE_URL}/rest/v1/inventario_perfileria?select=*&limit=2000'
req = urllib.request.Request(url, headers={
    'apikey': ANON_KEY,
    'Authorization': f'Bearer {ANON_KEY}'
})
with urllib.request.urlopen(req) as r:
    db_data = json.loads(r.read())

# Detectar columna de codigo en Supabase
db_codigos = set()
codigo_col = None
if db_data:
    print('Columnas en Supabase:', list(db_data[0].keys()))
    for key in db_data[0].keys():
        if 'cod' in key.lower() or 'codigo' in key.lower():
            codigo_col = key
            db_codigos = set(d.get(key, '') for d in db_data if d.get(key))
            print(f'Columna de codigo detectada: {key}')
            break
else:
    print('Supabase: tabla VACIA')
    # Intentar obtener columnas via HEAD request
    head_url = f'{SUPABASE_URL}/rest/v1/inventario_perfileria?select=*&limit=0'
    head_req = urllib.request.Request(head_url, headers={
        'apikey': ANON_KEY,
        'Authorization': f'Bearer {ANON_KEY}',
        'Prefer': 'count=exact'
    })
    with urllib.request.urlopen(head_req) as r:
        print('Headers:', dict(r.headers))

excel_codigos = set(por_codigo.keys())

print()
print('=' * 60)
print('RESUMEN COMPARATIVO')
print('=' * 60)
print(f'Excel - Total barras/registros:  {len(filled)}')
print(f'Excel - Codigos unicos:          {len(excel_codigos)}')
print(f'Supabase - Total registros:      {len(db_data)}')
print(f'Supabase - Codigos unicos:       {len(db_codigos)}')
print()
nuevos = excel_codigos - db_codigos
huerfanos = db_codigos - excel_codigos
comunes = excel_codigos & db_codigos
print(f'NUEVOS (en Excel, faltan en DB): {len(nuevos)}')
print(f'HUERFANOS (en DB, no en Excel):  {len(huerfanos)}')
print(f'COINCIDEN en ambos:              {len(comunes)}')
print()

# Tabla completa
print('=' * 60)
print('TABLA DE CODIGOS DEL EXCEL (stock en metros)')
print('=' * 60)
print(f"{'CODIGO':<12} {'DESCRIPCION':<40} {'COLOR':<15} {'BARRAS':>6} {'METROS':>8}  ESTADO")
print('-' * 100)

for cod in sorted(excel_codigos):
    items = por_codigo[cod]
    total_mm = sum(i['mm'] for i in items)
    total_m = round(total_mm / 1000, 3)
    desc = items[0]['desc'][:38]
    color = items[0]['color'][:13]
    estado = 'OK-COINCIDE' if cod in db_codigos else 'NUEVO-FALTA-EN-DB'
    print(f"{cod:<12} {desc:<40} {color:<15} {len(items):>6} {total_m:>8.3f}  {estado}")

if huerfanos:
    print()
    print('=' * 60)
    print('CODIGOS EN DB QUE NO ESTAN EN EL EXCEL:')
    for cod in sorted(huerfanos):
        print(f'  {cod}')
