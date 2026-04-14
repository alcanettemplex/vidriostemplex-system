
import urllib.request, json
from openpyxl import load_workbook
from collections import defaultdict

SUPABASE_URL = 'https://mvcmofhbjvxjnxefylsa.supabase.co'
SERVICE_KEY  = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im12Y21vZmhianZ4am54ZWZ5bHNhIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MjgzODMyNiwiZXhwIjoyMDg4NDE0MzI2fQ.MSKnNzTJPZa92XnIEFM8rjaiLKfjJSPBFMRgnJeeago'

HEADERS = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json'
}

def supabase_get(endpoint):
    req = urllib.request.Request(f'{SUPABASE_URL}/rest/v1/{endpoint}', headers=HEADERS)
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

# ── 1. Leer tabla Supabase ──────────────────────────────────────────────
print("Consultando Supabase...")
db_rows = supabase_get('inventario_perfileria?select=*&limit=2000')
print(f"  Registros en DB: {len(db_rows)}")
if db_rows:
    print(f"  Columnas: {list(db_rows[0].keys())}")

db_por_codigo = defaultdict(list)
for r in db_rows:
    cod = r.get('codigo', '').strip() if r.get('codigo') else 'SIN_CODIGO'
    db_por_codigo[cod].append(r)

db_codigos = set(db_por_codigo.keys())

# ── 2. Leer Excel hoja INV ──────────────────────────────────────────────
print("\nLeyendo Excel...")
wb = load_workbook(
    r'C:\Users\User\Desktop\vidrios-templex-system\documentation\INVENTARIO PERFILERIA ACTUAL 2026.xlsx',
    data_only=True
)
ws = wb['INV']
rows = list(ws.iter_rows(min_row=2, values_only=True))
filled = [r for r in rows if r[1] is not None and str(r[1]).strip() not in ('#N/A', '')]

excel_por_codigo = defaultdict(list)
for r in filled:
    cons, cod, desc, color, ref, mm, ubic = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
    cod = str(cod).strip()
    excel_por_codigo[cod].append({
        'cons': cons,
        'desc': str(desc).strip() if desc else '',
        'color': str(color).strip() if color else '',
        'ref': str(ref).strip() if ref else '',
        'mm': float(mm) if mm else 0,
        'ubic': str(ubic).strip() if ubic else ''
    })

excel_codigos = set(excel_por_codigo.keys())
print(f"  Barras en Excel:       {len(filled)}")
print(f"  Codigos unicos Excel:  {len(excel_codigos)}")

# ── 3. Comparación ─────────────────────────────────────────────────────
solo_excel  = excel_codigos - db_codigos
solo_db     = db_codigos    - excel_codigos
en_ambos    = excel_codigos & db_codigos

print(f"\n{'='*65}")
print("COMPARACION EXCEL vs SUPABASE")
print(f"{'='*65}")
print(f"  Registros actuales en DB:        {len(db_rows)}")
print(f"  Codigos unicos en DB:            {len(db_codigos)}")
print(f"  Codigos unicos en Excel:         {len(excel_codigos)}")
print(f"  [OK] Coinciden en ambos:         {len(en_ambos)}")
print(f"  [NUEVO] Solo en Excel (INSERTAR):{len(solo_excel)}")
print(f"  [HUERFANO] Solo en DB:           {len(solo_db)}")

# ── 4. Nuevos ───────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print("CODIGOS NUEVOS - EN EXCEL, FALTAN EN DB:")
print(f"{'='*65}")
if solo_excel:
    print(f"{'CODIGO':<12} {'DESCRIPCION':<38} {'COLOR':<14} {'BARRAS':>6} {'METROS':>8}")
    print('-'*82)
    for cod in sorted(solo_excel):
        items = excel_por_codigo[cod]
        total_m = round(sum(i['mm'] for i in items) / 1000, 3)
        print(f"{cod:<12} {items[0]['desc'][:37]:<38} {items[0]['color'][:13]:<14} {len(items):>6} {total_m:>8.3f}")
else:
    print("  Ninguno (todos los codigos del Excel ya estan en DB)")

# ── 5. Huerfanos ────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print("CODIGOS HUERFANOS - EN DB, NO EN EL EXCEL:")
print(f"{'='*65}")
if solo_db:
    print(f"{'CODIGO':<12} {'BARRAS_DB':>10} {'METROS_DB':>10}")
    print('-'*35)
    for cod in sorted(solo_db):
        items = db_por_codigo[cod]
        total_m = round(sum(float(i.get('mm', 0)) for i in items) / 1000, 3)
        print(f"{cod:<12} {len(items):>10} {total_m:>10.3f}")
else:
    print("  Ninguno")

# ── 6. Diferencias de stock en comunes ──────────────────────────────────
diffs = []
for cod in sorted(en_ambos):
    xls  = excel_por_codigo[cod]
    db   = db_por_codigo[cod]
    m_xls = round(sum(i['mm'] for i in xls) / 1000, 3)
    m_db  = round(sum(float(i.get('mm', 0)) for i in db) / 1000, 3)
    dif   = round(m_xls - m_db, 3)
    if abs(dif) > 0.01:
        diffs.append((cod, len(xls), m_xls, len(db), m_db, dif))

print(f"\n{'='*65}")
print(f"DIFERENCIAS DE STOCK EN CODIGOS COMUNES ({len(diffs)} con diferencia):")
print(f"{'='*65}")
if diffs:
    print(f"{'CODIGO':<12} {'B_XLS':>6} {'M_XLS':>8} {'B_DB':>6} {'M_DB':>8} {'DIFER':>9}")
    print('-'*55)
    for cod, bx, mx, bd, md, dif in diffs:
        print(f"{cod:<12} {bx:>6} {mx:>8.3f} {bd:>6} {md:>8.3f} {dif:>9.3f}")
else:
    print("  Todos los codigos comunes tienen el mismo stock en metros")
