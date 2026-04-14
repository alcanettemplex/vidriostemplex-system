
import urllib.request, json
from openpyxl import load_workbook
from collections import defaultdict

SUPABASE_URL = 'https://mvcmofhbjvxjnxefylsa.supabase.co'
ANON_KEY = 'sb_publishable_4P_mqyw7E_p2yPHuo8hCRA_BgNR0Wsm'

# --- Leer Excel hoja INV ---
wb = load_workbook(r'C:\Users\User\Desktop\vidrios-templex-system\documentation\INVENTARIO PERFILERIA ACTUAL 2026.xlsx', data_only=True)
ws = wb['INV']
rows = list(ws.iter_rows(min_row=2, values_only=True))
filled = [r for r in rows if r[1] is not None and str(r[1]) != '#N/A']

por_codigo = defaultdict(list)
for r in filled:
    cons, cod, desc, color, ref, mm, ubic = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
    por_codigo[cod].append({
        'cons': cons,
        'desc': str(desc) if desc else '',
        'color': str(color) if color else '',
        'ref': str(ref) if ref else '',
        'mm': mm or 0,
        'ubic': str(ubic) if ubic else ''
    })

# --- Consultar Supabase ---
url = f'{SUPABASE_URL}/rest/v1/inventario_perfileria?select=*&limit=2000'
req = urllib.request.Request(url, headers={
    'apikey': ANON_KEY,
    'Authorization': f'Bearer {ANON_KEY}'
})
with urllib.request.urlopen(req) as r:
    db_data = json.loads(r.read())

db_codigos = set()
if db_data:
    for key in db_data[0].keys():
        if 'cod' in key.lower():
            db_codigos = set(d.get(key, '') for d in db_data if d.get(key))
            break

excel_codigos = set(por_codigo.keys())
nuevos = excel_codigos - db_codigos
huerfanos = db_codigos - excel_codigos
comunes = excel_codigos & db_codigos

# --- Generar JSON para insercion ---
payload = []
for cod in sorted(excel_codigos):
    items = por_codigo[cod]
    total_mm = sum(i['mm'] for i in items)
    total_m = round(total_mm / 1000, 3)
    desc = items[0]['desc']
    color_principal = items[0]['color']
    ref_principal = items[0]['ref']
    # Ubicaciones unicas
    ubicaciones = list(set(i['ubic'] for i in items if i['ubic']))
    
    payload.append({
        'codigo_templex': cod,
        'descripcion': desc,
        'color': color_principal,
        'referencia': ref_principal,
        'stock_metros': total_m,
        'cantidad_barras': len(items),
        'ubicacion': ', '.join(sorted(ubicaciones)) if ubicaciones else None,
        'unidad_medida': 'metros'
    })

# Guardar JSON
with open(r'C:\Users\User\Desktop\vidrios-templex-system\documentation\inventario_para_supabase.json', 'w', encoding='utf-8') as f:
    json.dump(payload, f, ensure_ascii=False, indent=2)

print(f'JSON generado: {len(payload)} items')
print(f'Archivo: documentation/inventario_para_supabase.json')
print()
print('=== RESUMEN ===')
print(f'Excel - Total barras:           {len(filled)}')
print(f'Excel - Codigos unicos (items): {len(excel_codigos)}')
print(f'Supabase DB - Total registros:  {len(db_data)}')
print(f'')
print(f'Para insertar en DB:            {len(nuevos)} codigos nuevos')
print(f'Ya en DB (actualizar):          {len(comunes)} codigos')
print(f'En DB pero no en Excel:         {len(huerfanos)} codigos huerfanos')

# Primeros items del JSON
print()
print('=== MUESTRA DEL JSON A INSERTAR (primeros 5) ===')
for item in payload[:5]:
    print(json.dumps(item, ensure_ascii=False))
